"""Ingredient API routes."""

import io

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from sqlmodel import Session

from app.api.deps import OrgContext, get_current_user, get_org_context, get_session
from app.domain import IngredientService
from app.domain.category_service import CategoryService
from app.domain.fmh_import_service import (
    FMHImportResult,
    import_buy_catalogue,
    import_ingredients,
)
from app.domain.ingredient_service import NOT_ORG_ADMIN_FOR_MOVE, UNIT_NOT_FOUND
from app.domain.storage_service import (
    StorageError,
    StorageService,
    is_storage_configured,
)
from app.models import (
    Category,
    FoodCategory,
    Ingredient,
    IngredientCreate,
    IngredientSource,
    IngredientUpdate,
    SupplierIngredientCreate,
    SupplierIngredientRead,
    SupplierIngredientUpdate,
    User,
)
from app.passport import access

router = APIRouter()


def _require_org_admin(session: Session, current_user: User) -> None:
    """403 unless the caller administers the ORGANISATION.

    Ingredients are org-wide master data — a bulk import rewrites the catalogue for every brand at
    once, so there is no single unit to scope it to and no brand-level `Manager` check to make.
    """
    if not access.is_org_admin(session, current_user.id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only organisation administrators can perform this action",
        )


@router.post("", response_model=Ingredient, status_code=status.HTTP_201_CREATED)
def create_ingredient(
    data: IngredientCreate,
    session: Session = Depends(get_session),
    org: OrgContext = Depends(get_org_context),
):
    """Create a new ingredient."""
    service = IngredientService(session, org.organization_id)
    return service.create_ingredient(data)


@router.get("")
def list_ingredients(
    active_only: bool = True,
    category: FoodCategory | None = None,
    source: IngredientSource | None = None,
    master_only: bool = False,
    page_number: int = Query(default=1, ge=1),
    page_size: int = Query(default=30, ge=1, le=100),
    search: str | None = Query(default=None),
    category_ids: str | None = Query(default=None),
    units: str | None = Query(default=None),
    allergen_ids: str | None = Query(default=None),
    is_halal: str | None = Query(default=None),
    sort_by: str | None = Query(default=None),
    session: Session = Depends(get_session),
    org: OrgContext = Depends(get_org_context),
):
    """List all ingredients with optional filters."""
    from app.models.pagination import PaginatedResponse

    parsed_category_ids = [int(x) for x in category_ids.split(",")] if category_ids else None
    parsed_units = units.split(",") if units else None
    parsed_allergen_ids = [int(x) for x in allergen_ids.split(",")] if allergen_ids else None
    parsed_is_halal = [x.strip().lower() == "true" for x in is_halal.split(",")] if is_halal else None

    service = IngredientService(session, org.organization_id)
    offset = (page_number - 1) * page_size
    filter_kwargs = dict(active_only=active_only, category=category, source=source, master_only=master_only, search=search,
                         category_ids=parsed_category_ids, units=parsed_units, allergen_ids=parsed_allergen_ids, is_halal=parsed_is_halal,
                         sort_by=sort_by)
    items, total = service.list_paginated_with_count(offset=offset, limit=page_size, **filter_kwargs)
    return PaginatedResponse.create(items=items, total_count=total, page_number=page_number, page_size=page_size)


@router.post("/fmh-import", response_model=FMHImportResult)
async def import_ingredients_fmh(
    products_file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Import FMH categories, ingredients, and supplier links. Org administrators only."""
    _require_org_admin(session, current_user)
    if not (products_file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"File '{products_file.filename}' must be an .xlsx file",
        )
    products_wb = openpyxl.load_workbook(io.BytesIO(await products_file.read()), read_only=True, data_only=True)
    try:
        return import_ingredients(session, products_wb)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.post("/buy-catalogue-import", response_model=FMHImportResult)
async def import_ingredients_buy_catalogue(
    products_file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Import Buy Catalogue XLSX (single-sheet, inline supplier + SKU). Org administrators only."""
    _require_org_admin(session, current_user)
    if not (products_file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"File '{products_file.filename}' must be an .xlsx file",
        )
    wb = openpyxl.load_workbook(io.BytesIO(await products_file.read()), read_only=True, data_only=True)
    try:
        return import_buy_catalogue(session, wb)
    except (KeyError, ValueError) as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/buy-catalogue-template")
def download_buy_catalogue_template() -> Response:
    """Return a blank Buy Catalogue XLSX template with header row and one example row."""
    import io as _io

    import openpyxl as _openpyxl

    wb_out = _openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "EXPORT_BUY_CATALOGUE"
    headers = [
        "Rule Name", "Branch Name", "Product Name", "Supplier Name",
        "Sku", "Category Name", "Uom", "Unit", "Price", "Currency", "Packaging Note",
    ]
    examples = [
        ["Buy Catalogue", "DINING - CURLYS", "(EX001) Example Chicken (PKT)", "Chicken Supplier", "SUPP-FD-MEAT-000001", "Meat", "PKT", 1, 8.50, "SGD", "1KG / PKT"],
        ["Buy Catalogue", "DINING - CURLYS", "(EX002) Olive Oil (BTL)", "Oil Supplier", "SUPP-FD-OIL-000002", "Oil & Fat", "BTL", 1, 15.00, "SGD", "750ML / BTL"],
        ["Buy Catalogue", "DINING - CURLYS", "(EX003) Fresh Milk (LTR)", "Dairy Supplier", "SUPP-FD-DAIRY-000003", "Dairy", "LTR", 1, 3.20, "SGD", "1LTR / PKT"],
        ["Buy Catalogue", "DINING - CURLYS", "(EX004) Dinner Plate (PC)", "Tableware Supplier", "SUPP-EQ-WARE-000004", "Equipment", "PC", 1, 2.50, "SGD", "1PC / PC"],
        ["Buy Catalogue", "DINING - CURLYS", "(EX005) Rock Salt (KG)", "Seasoning Supplier", "SUPP-FD-SEAS-000005", "Seasoning", "KG", 1, 1.80, "SGD", "1KG / KG"],
    ]
    ws.append(headers)
    for row in examples:
        ws.append(row)
    buf = _io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type=_XLSX_CONTENT_TYPE,
        headers={"Content-Disposition": 'attachment; filename="BuyCatalogue_template.xlsx"'},
    )


@router.get("/fmh-sample-items")
async def download_fmh_sample_items() -> Response:
    """Download the FMH sample product list XLSX template."""
    if not is_storage_configured():
        raise HTTPException(status_code=503, detail="Storage not configured")
    try:
        data = await StorageService().download_fmh_sample("ProductList_sample.xlsx")
    except StorageError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return Response(
        content=data,
        media_type=_XLSX_CONTENT_TYPE,
        headers={"Content-Disposition": 'attachment; filename="ProductList_sample.xlsx"'},
    )


@router.get("/categories", response_model=list[Category])
def list_categories(
    session: Session = Depends(get_session),
    org: OrgContext = Depends(get_org_context),
):
    """List all available ingredient categories from the database."""
    service = CategoryService(session, org.organization_id)
    return service.list_categories(active_only=True)


@router.get("/{ingredient_id}", response_model=Ingredient)
def get_ingredient(
    ingredient_id: int,
    session: Session = Depends(get_session),
    org: OrgContext = Depends(get_org_context),
):
    """Get an ingredient by ID."""
    service = IngredientService(session, org.organization_id)
    ingredient = service.get_ingredient(ingredient_id)
    if not ingredient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ingredient not found",
        )
    return ingredient


@router.get("/{ingredient_id}/variants", response_model=list[Ingredient])
def get_variants(
    ingredient_id: int,
    session: Session = Depends(get_session),
    org: OrgContext = Depends(get_org_context),
):
    """Get all variant ingredients linked to a master ingredient."""
    service = IngredientService(session, org.organization_id)

    # Verify the ingredient exists
    ingredient = service.get_ingredient(ingredient_id)
    if not ingredient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ingredient not found",
        )

    return service.get_variants(ingredient_id)


@router.patch("/{ingredient_id}", response_model=Ingredient)
def update_ingredient(
    ingredient_id: int,
    data: IngredientUpdate,
    session: Session = Depends(get_session),
    org: OrgContext = Depends(get_org_context),
):
    """Update an ingredient."""
    service = IngredientService(session, org.organization_id)
    ingredient = service.update_ingredient(ingredient_id, data)
    if not ingredient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ingredient not found",
        )
    return ingredient


@router.patch("/{ingredient_id}/deactivate", response_model=Ingredient)
def deactivate_ingredient(
    ingredient_id: int,
    session: Session = Depends(get_session),
    org: OrgContext = Depends(get_org_context),
):
    """Deactivate (soft-delete) an ingredient."""
    service = IngredientService(session, org.organization_id)
    ingredient = service.deactivate_ingredient(ingredient_id)
    if not ingredient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ingredient not found",
        )
    return ingredient


# -----------------------------------------------------------------------------
# Supplier Management Endpoints
# -----------------------------------------------------------------------------


@router.get("/{ingredient_id}/suppliers", response_model=list[SupplierIngredientRead])
def get_ingredient_suppliers(
    ingredient_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    org: OrgContext = Depends(get_org_context),
):
    """Get all suppliers for an ingredient, restricted to the units the caller can see."""
    service = IngredientService(session, org.organization_id)
    result = service.get_ingredient_suppliers(
        ingredient_id,
        subject=current_user.id,
    )
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ingredient not found",
        )
    return result


@router.post(
    "/{ingredient_id}/suppliers",
    response_model=SupplierIngredientRead,
    status_code=status.HTTP_201_CREATED,
)
def add_ingredient_supplier(
    ingredient_id: int,
    data: SupplierIngredientCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    org: OrgContext = Depends(get_org_context),
):
    """Add a supplier to an ingredient, at a unit the caller can actually see.

    `data.unit_id` is client-supplied, so it is checked against the caller's brands — otherwise
    anyone could attach pricing to a rival brand's unit.
    """
    # Ensure the path ingredient_id matches the body
    data.ingredient_id = ingredient_id
    service = IngredientService(session, org.organization_id)
    result = service.add_ingredient_supplier(ingredient_id, data, subject=current_user.id)
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ingredient or supplier not found",
        )
    if isinstance(result, str):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=result,
        )
    return result


@router.patch(
    "/{ingredient_id}/suppliers/{supplier_ingredient_id}",
    response_model=SupplierIngredientRead,
)
def update_ingredient_supplier(
    ingredient_id: int,
    supplier_ingredient_id: int,
    data: SupplierIngredientUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    org: OrgContext = Depends(get_org_context),
):
    """Update a supplier-ingredient link the caller can see.

    Scope is checked in the service against the link's current unit — and, when moving, the target
    unit too. The route previously checked authority ONLY when `unit_id` was present, so a PATCH
    without it reached the service unchecked and could rewrite any brand's pricing.

    Two different denials, deliberately different codes:
      - 404 when the link is not visible to the caller — the scoped read would never have shown it,
        so a write must not confirm it exists either.
      - 403 when the caller CAN see the link but may not re-home it. Re-homing stays
        org-admin-only; the row is not a secret from them, the operation is simply refused.
    """
    service = IngredientService(session, org.organization_id)
    result = service.update_ingredient_supplier(
        supplier_ingredient_id, data, subject=current_user.id
    )
    if result == NOT_ORG_ADMIN_FOR_MOVE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=result)
    if result == UNIT_NOT_FOUND:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=result)
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Supplier-ingredient link not found",
        )
    return result


@router.delete(
    "/{ingredient_id}/suppliers/{supplier_ingredient_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
def remove_ingredient_supplier(
    ingredient_id: int,
    supplier_ingredient_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    org: OrgContext = Depends(get_org_context),
):
    """Remove a supplier from an ingredient, if the caller can see the link.

    404 rather than 403 for a link at another brand's unit: its existence is not the caller's to
    learn, and the scoped read would never have shown it to them.
    """
    service = IngredientService(session, org.organization_id)
    success = service.remove_ingredient_supplier(
        supplier_ingredient_id, subject=current_user.id
    )
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Supplier-ingredient link not found",
        )


@router.get(
    "/{ingredient_id}/suppliers/preferred",
    response_model=SupplierIngredientRead | None,
)
def get_preferred_supplier(
    ingredient_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    org: OrgContext = Depends(get_org_context),
):
    """Get the preferred supplier for an ingredient, restricted to the units the caller can see."""
    service = IngredientService(session, org.organization_id)
    ingredient = service.get_ingredient(ingredient_id)
    if not ingredient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ingredient not found",
        )
    return service.get_preferred_supplier(
        ingredient_id,
        subject=current_user.id,
    )
